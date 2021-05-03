
from win32com.client import Dispatch
import datetime
import win32com.client as win32
from openpyxl import Workbook
import json
import os
from openpyxl.styles import PatternFill, Font

class OutlookDispatcher:
    def __init__(self):

        self.employee_dic = {}
        with open('user_settings.json', 'r') as myfile:
            data = myfile.read()
        self.json_obj = json.loads(data)

        self.outlook = Dispatch("Outlook.Application").GetNamespace("MAPI")
        self.OUTLOOK_FORMAT = '%m/%d/%Y %H:%M'
        self.category = self.json_obj['categories']
        self.appointments = []
        self.calendar = self.json_obj['calendar']
        self.set_time_interval()
        self.delete_excel_file()

        self.workbook = Workbook()
        self.worksheet = self.workbook["Sheet"]

    def write_to_json(self):
        with open("user_settings.json", "w") as jsonFile:
            json.dump(self.json_obj, jsonFile)

    def give_file_name(self):
        user_name = str(self.outlook.GetDefaultFolder(9).Parent).split('@')[0]
        timePeriod = str(self.begin.strftime("%Y-%m-%d") + "_" + self.end.strftime("%Y-%m-%d"))
        subject = str("OB")
        return str(subject + "_" + timePeriod + "_" + user_name + ".xlsx")

    def get_appointments(self):
        for calendar in self.calendar:
            try:
                if str(calendar) == 'YourMainCalendar':
                    restricted_items = self.outlook.GetDefaultFolder(9).Items
                    restricted_items.Sort("[Start]")
                    restricted_items.IncludeRecurrences = True
                    self.filter_restrictions(restricted_items)
                else:
                    restricted_items = self.outlook.GetDefaultFolder(9).Folders[calendar].Items
                    restricted_items.Sort("[Start]")
                    restricted_items.IncludeRecurrences = True
                    self.filter_restrictions(restricted_items)


            except Exception:
                pass
        self.sort_com_object()

    def sort_com_object(self):
        for i in range(self.appointments.__len__()):
            small = datetime.datetime.strptime(self.appointments.__getitem__(i).Start.Format('%Y/%m/%d'),
                                               '%Y/%m/%d').date()
            for j in range(self.appointments.__len__()):
                small2 = datetime.datetime.strptime(self.appointments.__getitem__(j).Start.Format('%Y/%m/%d'),
                                                    '%Y/%m/%d').date()
                if small2 > small:
                    small = datetime.datetime.strptime(self.appointments.__getitem__(j).Start.Format('%Y/%m/%d'),
                                                    '%Y/%m/%d').date()
                    temp = self.appointments.__getitem__(i)
                    self.appointments.__setitem__(i, self.appointments.__getitem__(j))
                    self.appointments.__setitem__(j, temp)

    def set_time_interval(self):
        today = datetime.date.today()
        monday = today - datetime.timedelta(days=today.weekday())
        self.begin = monday
        self.end = monday + datetime.timedelta(days=6)

    def filter_restrictions(self, restricted_items):
        restriction = "[Start] >= '" + self.begin.strftime("%m/%d/%Y") + "' AND [End] <= '" + self.end.strftime(
            "%m/%d/%Y") + "'"
        restricted_items = restricted_items.Restrict(restriction)
        self.add_filtered_appointments(restricted_items)

    def add_filtered_appointments(self, restricted_items):
        for item in restricted_items:
            if item.Categories == self.category[0] or item.Categories == self.category[1]:
                self.appointments.append(item)

    def user_interaction(self):
        userinput = ""
        while userinput != 0:
            print("\n")
            print("Current settings:")
            print("Time period: " + self.begin.strftime("%Y/%m/%d") + " - " + self.end.strftime(
                "%Y/%m/%d"))
            print("Category: " + self.category[0] + ", " + self.category[1])
            print("Calendar: " + self.calendar[0] + ", " + self.calendar[1])
            print("-------------------------------------------")
            print("Press 1 to change the time period")
            print("Press 2 to change category")
            print("Press 3 to change calendar")
            print("Press 4 to produce excel sheet")
            print("Press 0 to exit")
            userinput = input("Your input: ")
            if userinput == str(1):
                self.change_time_period()

            if userinput == str(2):
                self.change_category()

            if userinput == str(3):
                self.change_calendar()

            if userinput == str(4):
                self.produce_excel_sheet()

            if userinput == str(0):
                exit()






    def change_time_period(self):
        print("example: 2021.04.30")
        userinput = input("begin: ")
        split = userinput.split('.')
        self.begin = datetime.date(int(split[0]), int(split[1]), int(split[2]))
        userinput = input("end: ")
        split = userinput.split('.')
        self.end = datetime.date(int(split[0]), int(split[1]), int(split[2]))
        print("Time period was changed successfully")

    def change_category(self):
        self.category[0] = str(input("choose new category 1: "))
        self.category[1] = str(input("choose new category 2: "))
        self.json_obj['categories'][0] = self.category[0]
        self.json_obj['categories'][1] = self.category[1]
        self.write_to_json()
        print("category was changed successfully!")

    def change_calendar(self):
        print("Your Main Calendar will be always selected.")
        self.calendar[1] = str(input("choose a calendar: "))
        self.json_obj['calendar'][1] = self.calendar[1]
        self.write_to_json()
        print("calendar was changed successfully!")

    def write_excel(self):
        self.excel_header()
        self.add_appointments()
        self.workbook.save(filename=self.json_obj['file_name'])

    def excel_header(self):
        self.get_emp()
        self.create_employee_dic()
        header = ['Calendar Week', 'Date', 'Topic', 'Description', 'Team'] + list(self.employeeSet) + ["Σ"]
        for i in range(len(header)):
            self.worksheet.cell(row=1, column=i + 1, value=header[i])
            self.worksheet.cell(row=1, column=i + 1).fill = PatternFill(start_color='14e31f', end_color='14e31f',
                                                                        fill_type="solid")
            self.worksheet.cell(row=1, column=i + 1).font = Font(bold=True, size=10)

    def get_emp(self):
        employee_set = set()
        for appointmentItem in self.appointments:
            for i in range(appointmentItem.recipients.Count):
                if not str(appointmentItem.recipients[i]).__contains__("bmw"):
                    employee_set.add(str(appointmentItem.recipients[i]))
        self.employeeSet = sorted(employee_set)

    def create_employee_dic(self):
        for i in range(self.employeeSet.__len__()):
            self.employee_dic.__setitem__(self.employeeSet.__getitem__(i), i + 6)

    def set_row_width(self):
        self.worksheet.column_dimensions['A'].width = 5
        self.worksheet.column_dimensions['B'].width = 12
        self.worksheet.column_dimensions['C'].width = 40
        self.worksheet.column_dimensions['D'].width = 40
        self.worksheet.column_dimensions['E'].width = 12

    def add_appointments(self):
        row = 2
        self.set_row_width()

        for appointmentItem in self.appointments:
            self.worksheet.cell(row=row, column=1, value=
            datetime.datetime.strptime(appointmentItem.Start.Format('%Y/%m/%d'), '%Y/%m/%d').date().isocalendar()[1])
            self.worksheet.cell(row=row, column=2,
                                value=datetime.datetime.strptime(appointmentItem.Start.Format('%Y/%m/%d'),
                                                                 '%Y/%m/%d').date())
            self.worksheet.cell(row=row, column=3, value=appointmentItem.Subject)
            if str(appointmentItem.categories) == self.json_obj['categories'][1]:
                for i in range(self.employeeSet.__len__() + 6):
                    self.worksheet.cell(row=row, column=i + 1).fill = PatternFill(start_color='fff67f',
                                                                                  end_color='fff67f',
                                                                                  fill_type="solid")

            self.worksheet.cell(row=row, column=5, value=appointmentItem.categories)
            sum_duration = 0
            for emp in self.employeeSet:
                self.worksheet.cell(row=row, column=self.employee_dic.get(emp),
                                    value=0)

            for i in range(appointmentItem.recipients.Count):
                if self.employee_dic.keys().__contains__(str(appointmentItem.recipients[i])):

                    if appointmentItem.ResponseStatus != 4:
                        self.worksheet.cell(row=row, column=self.employee_dic.get(str(appointmentItem.recipients[i])),
                                            value=round(appointmentItem.Duration / 60,2))
                        sum_duration += round(appointmentItem.Duration / 60,2)
                    else:
                        self.worksheet.cell(row=row, column=self.employee_dic.get(str(appointmentItem.recipients[i])),
                                            value=0)
            self.worksheet.cell(row=row, column=self.employee_dic.__len__() + 6, value=sum_duration)
            row += 1

    def produce_excel_sheet(self):
        self.get_appointments()
        self.workbook.title = self.json_obj['file_name']
        self.json_obj['file_name'] = self.give_file_name()
        print("excel sheet will be finished shortly!")
        self.write_to_json()
        self.write_excel()

        excel = win32.gencache.EnsureDispatch('Excel.Application')

        excel.Workbooks.Open(Filename=os.path.realpath(self.json_obj['file_name']))
        excel.Visible = True
        exit()

    def delete_excel_file(self):
        if self.json_obj['file_name'] != "":
            os.remove(self.json_obj['file_name'])
            self.json_obj['file_name'] = ""
            self.write_to_json()


def main():
    a = OutlookDispatcher()
    a.user_interaction()


if __name__ == "__main__":
    main()

